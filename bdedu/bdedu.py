import openpyxl
import re
import json
import csv
import time

import os
projectPath = os.path.normpath(os.getcwd() + os.sep + os.pardir)

import threading
from concurrent.futures import ThreadPoolExecutor



finalData = dict()

class bdEduDataCollector():
    # data source
    # https://banbeis.portal.gov.bd/

    def __init__(self):
        self.finalData = dict()
        self.finalData["source"] = "https://banbeis.portal.gov.bd/"
        self.finalData["description"] = "Information of each educational institute in Bangladesh"
        self.finalData["creation time"] = int(time.time())
        self.finalData['organization'] = dict()

    def saveData(self):
        with open('finalData.json', 'w') as f:
            json.dump(self.finalData, f, indent=2)
        print("\ntotal data loaded:", len(self.finalData['organization']))


    def readSchool(self):
        # data source
        # http://banbeis.portal.gov.bd/sites/default/files/files/banbeis.portal.gov.bd/npfblock/school.xlsx

        sheets = openpyxl.load_workbook('school.xlsx')
        sheet = sheets['Sheet1']
        rows = sheet.rows

        header = [str(cell.value) for cell in next(rows)]

        p = re.compile(r"[^a-zA-Z0-9]")
        for i in range(len(header)):
            x = p.split(header[i])
            x = ' '.join([j.lower().strip() for j in x if x != ''])
            header[i] = x

        for row in rows:
            rowData = [cell.value for cell in row]
            school = dict()
            for i in range(len(rowData)):
                school[header[i]] = rowData[i]

            d = dict()
            address = ""
            if school["vill road"] != None:
                if len(address):
                    address = address + ', '
                address = address + school["vill road"]
            if school["thana name"] != None:
                if len(address):
                    address = address + ', '
                address = address + school["thana name"]
            if school["district name"] != None:
                if len(address):
                    address = address + ', '
                address = address + school["district name"]
            if school["division name"] != None:
                if len(address):
                    address = address + ', '
                address = address + school["division name"]

            d['name'] = ' '.join([x.strip() for x in school["institute name"].split() if x.strip() != ''])
            d['address'] = address
            d['country'] = 'Bangladesh'
            d['tel'] = school['mobile']

            self.finalData['organization'][d['name'].lower()] = d

        self.saveData()
        print("\nschool data fatched!")

    def readCollege(self):
        # data source
        # http://banbeis.portal.gov.bd/sites/default/files/files/banbeis.portal.gov.bd/npfblock/college.xlsx

        sheets = openpyxl.load_workbook('college.xlsx')
        sheet = sheets['Sheet1']
        rows = sheet.rows

        header = [str(cell.value) for cell in next(rows)]

        p = re.compile(r"[^a-zA-Z0-9]")
        for i in range(len(header)):
            x = p.split(header[i])
            x = ' '.join([j.lower().strip() for j in x if x != ''])
            header[i] = x

        for row in rows:
            rowData = [cell.value for cell in row]
            school = dict()
            for i in range(len(rowData)):
                school[header[i]] = rowData[i]

            d = dict()
            address = ""
            if school["vill road"] != None:
                if len(address):
                    address = address + ', '
                address = address + school["vill road"]
            if school["thana name"] != None:
                if len(address):
                    address = address + ', '
                address = address + school["thana name"]
            if school["district name"] != None:
                if len(address):
                    address = address + ', '
                address = address + school["district name"]
            if school["division name"] != None:
                if len(address):
                    address = address + ', '
                address = address + school["division name"]

            d['name'] = ' '.join([x.strip() for x in school["institute name"].split() if x.strip() != ''])
            d['address'] = address
            d['country'] = 'Bangladesh'
            d['tel'] = school['mobile']

            self.finalData['organization'][d['name'].lower()] = d

        self.saveData()
        print("\nCollege data fatched!")


    def readMadrasa(self):
        # data source
        # http://banbeis.portal.gov.bd/sites/default/files/files/banbeis.portal.gov.bd/npfblock/madrasa.xlsx

        sheets = openpyxl.load_workbook('madrasa.xlsx')
        sheet = sheets['Sheet1']
        rows = sheet.rows

        header = [str(cell.value) for cell in next(rows)]

        p = re.compile(r"[^a-zA-Z0-9]")
        for i in range(len(header)):
            x = p.split(header[i])
            x = ' '.join([j.lower().strip() for j in x if x != ''])
            header[i] = x

        for row in rows:
            rowData = [cell.value for cell in row]
            school = dict()
            for i in range(len(rowData)):
                school[header[i]] = rowData[i]

            d = dict()
            address = ""
            if school["vill road"] != None:
                if len(address):
                    address = address + ', '
                address = address + school["vill road"]
            if school["thana name"] != None:
                if len(address):
                    address = address + ', '
                address = address + school["thana name"]
            if school["district name"] != None:
                if len(address):
                    address = address + ', '
                address = address + school["district name"]
            if school["division name"] != None:
                if len(address):
                    address = address + ', '
                address = address + school["division name"]

            d['name'] = ' '.join([x.strip() for x in school["institute name"].split() if x.strip() != ''])
            d['address'] = address
            d['country'] = 'Bangladesh'
            d['tel'] = school['mobile']

            self.finalData['organization'][d['name'].lower()] = d

        self.saveData()
        print("\nMadrasa data fatched!")


    def readSchoolCollege(self):
        # data source
        # http://banbeis.portal.gov.bd/sites/default/files/files/banbeis.portal.gov.bd/npfblock/School%26college.xlsx

        sheets = openpyxl.load_workbook('schoolcollege.xlsx')
        sheet = sheets['Sheet1']
        rows = sheet.rows

        header = [str(cell.value) for cell in next(rows)]

        p = re.compile(r"[^a-zA-Z0-9]")
        for i in range(len(header)):
            x = p.split(header[i])
            x = ' '.join([j.lower().strip() for j in x if x != ''])
            header[i] = x

        for row in rows:
            rowData = [cell.value for cell in row]
            school = dict()
            for i in range(len(rowData)):
                school[header[i]] = rowData[i]

            d = dict()
            address = ""
            if school["vill road"] != None:
                if len(address):
                    address = address + ', '
                address = address + school["vill road"]
            if school["thana name"] != None:
                if len(address):
                    address = address + ', '
                address = address + school["thana name"]
            if school["district name"] != None:
                if len(address):
                    address = address + ', '
                address = address + school["district name"]
            if school["division name"] != None:
                if len(address):
                    address = address + ', '
                address = address + school["division name"]

            d['name'] = ' '.join([x.strip() for x in school["institute name"].split() if x.strip() != ''])
            d['address'] = address
            d['country'] = 'Bangladesh'
            d['tel'] = school['mobile']

            self.finalData['organization'][d['name'].lower()] = d

        self.saveData()
        print("\nSchool and College data fatched!")


    def readTechnical(self):

        sheets = openpyxl.load_workbook('technical.xlsx')
        sheet = sheets['Sheet 1']
        rows = sheet.rows

        header = [str(cell.value) for cell in next(rows)]

        p = re.compile(r"[^a-zA-Z0-9]")
        for i in range(len(header)):
            x = p.split(header[i])
            x = ' '.join([j.lower().strip() for j in x if x != ''])
            header[i] = x

        for row in rows:
            rowData = [cell.value for cell in row]
            school = dict()
            for i in range(len(rowData)):
                school[header[i]] = rowData[i]

            d = dict()
            address = ""
            if school["upazila thana"] != None:
                if len(address):
                    address = address + ', '
                address = address + school["upazila thana"]
            if school["district"] != None:
                if len(address):
                    address = address + ', '
                address = address + school["district"]
            if school["division"] != None:
                if len(address):
                    address = address + ', '
                address = address + school["division"]

            d['name'] = ' '.join([x.strip() for x in school["name"].split() if x.strip() != ''])
            d['address'] = address
            d['country'] = 'Bangladesh'
            d['tel'] = school["mobile"]

            self.finalData['organization'][d['name'].lower()] = d

        self.saveData()
        print("\nTechnical college data fatched!")

def main():
    bdedu = bdEduDataCollector()

    with ThreadPoolExecutor(max_workers=5) as executor:
        executor.submit(bdedu.readCollege())
        executor.submit(bdedu.readSchool())
        executor.submit(bdedu.readSchoolCollege())
        executor.submit(bdedu.readMadrasa())
        executor.submit(bdedu.readTechnical())


if __name__ == '__main__':
    main()
